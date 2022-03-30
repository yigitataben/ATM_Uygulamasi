from re import A
import time
import sys
from openpyxl import load_workbook
import pandas as pd
import os
import getpass
import pwinput
os.getcwd()

excel_data_frame = load_workbook(filename="ATM_Uygulamasi_Database.xlsx")
worksheet = excel_data_frame.active

# Kontrol değişkenleri
is_only_password_wrong = False
is_only_password_wrong_index = 1
is_entry_true = False
is_entry_true_index = 1

# Komutlar
sys_firewall = worksheet.cell(row=is_entry_true_index,column=4).value
entree = 3

def menu():
    while True:
            print("""
******************************************************
Akademi Kurumsal Bankacılık Sistemi'ne Hoş Geldiniz...
******************************************************
""")
            islem = input("""
*****************************************************************
Lütfen yapmak istediğiniz işleme karşılık gelen numarayı giriniz:

1 - "Kendi Hesabına Para Yatırma"
2 - "Kendi Hesabından Para Çekme"
3 - "Bakiye Sorgulama"
0 - "Çıkış"

*****************************************************************
""")
            match islem:
                case "1":
                    print("Şuanki bakiyeniz: {}₺'dir.".format(worksheet.cell(row=is_entry_true_index,column=3).value))
                    try:
                        bakiye_yatirma = int(input("Hesabınıza yatırmak istediğiniz tutarı ₺ cinsinden giriniz."))
                        if (bakiye_yatirma > 0):
                            worksheet.cell(row=is_entry_true_index,column=3).value += bakiye_yatirma
                            excel_data_frame.save("ATM_Uygulamasi_Database.xlsx")
                            print("Güncel bakiyeniz: {}₺'dir.".format(worksheet.cell(row=is_entry_true_index,column=3).value))

                        else:
                            print("Yatırmak istediğiniz tutar 0'dan büyük olmalıdır.")
                        
                    except:
                        print("Yatırmak istediğiniz tutar harf içermemelidir.")

                case "2":
                    print("Şuanki bakiyeniz: {}₺'dir.".format(worksheet.cell(row=is_entry_true_index,column=3).value))
                    try:
                        bakiye_cekme = int(input("Hesabınızdan çekmek istediğiniz tutarı ₺ cinsinden giriniz."))
                        if (bakiye_cekme <= worksheet.cell(row=is_entry_true_index,column=3).value):
                            worksheet.cell(row=is_entry_true_index,column=3).value -= bakiye_cekme
                            excel_data_frame.save("ATM_Uygulamasi_Database.xlsx")
                            print("Güncel bakiyeniz: {}₺'dir.".format(worksheet.cell(row=is_entry_true_index,column=3).value))

                        else:
                            print("Hesabınızda bu kadar para yok.")

                    except:
                        print("Çekmek istediğiniz tutar harf içermemelidir.")

                case "3":
                    print("Hesabınızdaki güncel tutar: {}₺'dir.".format(worksheet.cell(row=is_entry_true_index,column=3).value))

                case "0":
                    t = 3
                    while t:
                        mins, secs = divmod(t, 60)
                        timer = '{:02d}:{:02d}'.format(mins, secs)
                        print ("Güvenli çıkış yapılıyor...",timer)
                        time.sleep(1)
                        t -= 1
                    sys.exit()   

                case _:

                    print("Geçersiz işlem numarası. Lütfen geçerli bir numara giriniz...")


while (entree > 0):
    card_number = input("Lütfen Kart Numaranızı Giriniz:")
    password = pwinput.pwinput("Lütfen Şifrenizi Giriniz:")
    card_number_index = 1

    for i in range(1, worksheet.max_row + 1):
        if(card_number == str(worksheet.cell(row=i, column=1).value)):
            card_number_index = i
            break

    if(card_number == str(worksheet.cell(row=card_number_index, column=1).value) and password != str(worksheet.cell(row=i, column=2).value)):
            entree -= 1
            print ("Şifreniz yanlış. Kalan giriş hakkınız: ",entree)
            if (entree == 2):
                answer = input("Şifrenizi değiştirmek ister misiniz? EVET ya da HAYIR diyebilirsinz.").upper()
                
                while (entree > 0):
                    if (answer == "EVET"):
                        firewall = input("Lütfen anne kızlık soyadınızın ilk iki hanesini giriniz.").upper()

                        if (firewall == sys_firewall):
                            sys_password = input("Lütfen yeni şifrenizi giriniz.")
                            worksheet.cell(row=card_number_index, column=2).value = sys_password
                            excel_data_frame.save("ATM_Uygulamasi_Database.xlsx")
                            entree += 1
                            print("Şifre güvenle değiştirildi.")
                            break

                        if (firewall != sys_firewall):
                            print("Şifre değiştirme işlemi başarısız. Lütfen güvenlik sorusuna geçerli bir yanıt giriniz.")
                            entree -= 1

                    elif (answer == "HAYIR"):
                        continue

                    else:
                        print("Geçersiz komut. Çıkış yapılıyor.")
                        break

    elif(card_number != str(worksheet.cell(row=card_number_index, column=1).value) and password != str(worksheet.cell(row=i, column=2).value)):
        print ("Kart numaranız ve şifreniz yanlış, lütfen tekrar deneyiniz.")
        entree -= 1
        print ("Kalan giriş hakkınız: ",entree)
        continue

    elif(card_number == str(worksheet.cell(row=card_number_index, column=1).value) and password == str(worksheet.cell(row=i, column=2).value)):
        menu()

    else:
        print("Kart numarası bulunamadı.")
        entree -= 1
        t = 3
        while t:
            mins, secs = divmod(t, 60)
            timer = '{:02d}:{:02d}'.format(mins, secs)
            print ("Güvenli çıkış yapılıyor...",timer)
            time.sleep(1)
            t -= 1
            sys.exit()
        continue